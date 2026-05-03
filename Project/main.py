"""
Агрегатор вакансий и трекер поиска работы
Консольное меню приложения
"""

import json
import csv
import os
from datetime import datetime
from typing import Optional

from app.vacancy import Vacancy
from app.tracker import Application, JobTracker
from app.parser import JobParser
from app.analytics import Analytics


class JobTrackerApp:
    """Основной класс приложения"""

    DATA_DIR = "data"
    VACANCIES_FILE = os.path.join(DATA_DIR, "vacancies.json")
    LOG_FILE = os.path.join(DATA_DIR, "operation_log.csv")

    def __init__(self):
        self._vacancies: list[Vacancy] = []
        self._tracker = JobTracker()
        self._parser = JobParser()
        self._analytics: Optional[Analytics] = None
        self._load_data()

    def _log_operation(self, operation: str, details: str):
        """Записать операцию в лог CSV"""
        try:
            with open(self.LOG_FILE, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    datetime.now().isoformat(),
                    operation,
                    details,
                    'user'
                ])
        except Exception as e:
            print(f"⚠️ Ошибка записи лога: {e}")

    def _load_data(self):
        """Загрузить данные из JSON файлов"""
        try:
            if os.path.exists(self.VACANCIES_FILE):
                with open(self.VACANCIES_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                # Загрузка вакансий
                for v_data in data.get('vacancies', []):
                    vacancy = Vacancy.from_dict(v_data)
                    self._vacancies.append(vacancy)

                # Загрузка откликов
                for a_data in data.get('applications', []):
                    application = Application.from_dict(a_data)
                    self._tracker.add_application(
                        application.vacancy,
                        application.notes
                    )
                    # Восстанавливаем статус и даты
                    apps = self._tracker.get_all_applications()
                    if apps:
                        last_app = apps[-1]
                        last_app._status = application.status
                        last_app._applied_date = application.applied_date
                        last_app._response_date = application.response_date

                self._analytics = Analytics(self._vacancies, self._tracker.get_all_applications())
                self._log_operation("LOAD_DATA", f"Loaded {len(self._vacancies)} vacancies, {len(self._tracker)} applications")
                print(f"✅ Загружено {len(self._vacancies)} вакансий и {len(self._tracker)} откликов")
            else:
                print("⚠️ Файл данных не найден. Создаём новую базу.")
                self._save_data()
        except Exception as e:
            print(f"❌ Ошибка загрузки данных: {e}")
            self._vacancies = []
            self._tracker = JobTracker()

    def _save_data(self):
        """Сохранить данные в JSON файлы"""
        try:
            os.makedirs(self.DATA_DIR, exist_ok=True)

            data = {
                'vacancies': [v.to_dict() for v in self._vacancies],
                'applications': [app.to_dict() for app in self._tracker.get_all_applications()]
            }

            with open(self.VACANCIES_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            self._log_operation("SAVE_DATA", f"Saved {len(self._vacancies)} vacancies, {len(self._tracker)} applications")
            print("✅ Данные сохранены")
        except Exception as e:
            print(f"❌ Ошибка сохранения данных: {e}")

    def _show_menu(self) -> str:
        """Показать главное меню"""
        menu = """
╔═══════════════════════════════════════════════════════╗
║         АГРЕГАТОР ВАКАНСИЙ И ТРЕКЕР ПОИСКА РАБОТЫ     ║
╠═══════════════════════════════════════════════════════╣
║  1. Показать все вакансии                             ║
║  2. Фильтровать вакансии                              ║
║  3. Парсить новые вакансии (hh.ru)                    ║
║  4. Подать отклик на вакансию                         ║
║  5. Показать мои отклики                              ║
║  6. Обновить статус отклика                           ║
║  7. Аналитика рынка                                   ║
║  8. Экспорт в Excel                                   ║
║  9. Сохранить данные                                  ║
║  0. Выход                                             ║
╚═══════════════════════════════════════════════════════╝
"""
        return menu

    def _filter_vacancies(self):
        """Фильтрация вакансий с использованием lambda"""
        print("\n=== ФИЛЬТР ВАКАНСИЙ ===")

        # Фильтр по зарплате
        min_salary = input("Минимальная зарплата (Enter - пропустить): ").strip()
        if min_salary:
            try:
                min_salary = float(min_salary)
                self._vacancies = list(filter(lambda v: v.get_average_salary() >= min_salary, self._vacancies))
                print(f"✅ Отфильтровано по зарплате от {min_salary}")
            except ValueError:
                print("⚠️ Некорректное значение зарплаты")

        # Фильтр по городу
        city = input("Город (Enter - пропустить): ").strip()
        if city:
            self._vacancies = list(filter(lambda v: city.lower() in v.city.lower(), self._vacancies))
            print(f"✅ Отфильтровано по городу '{city}'")

        # Фильтр по опыту
        print(f"Доступные уровни опыта: {Vacancy.VALID_EXPERIENCE_LEVELS}")
        experience = input("Уровень опыта (Enter - пропустить): ").strip()
        if experience and experience in Vacancy.VALID_EXPERIENCE_LEVELS:
            self._vacancies = list(filter(lambda v: v.experience_level == experience, self._vacancies))
            print(f"✅ Отфильтровано по опыту '{experience}'")

        # Фильтр по ключевым словам
        keywords = input("Ключевые слова (через запятую, Enter - пропустить): ").strip()
        if keywords:
            kw_list = [k.strip() for k in keywords.split(',')]
            self._vacancies = list(filter(lambda v: v.matches_keywords(kw_list), self._vacancies))
            print(f"✅ Отфильтровано по ключевым словам: {keywords}")

        self._show_vacancies()

    def _show_vacancies(self):
        """Показать список вакансий"""
        print(f"\n=== ВАКАНСИИ ({len(self._vacancies)}) ===")

        if not self._vacancies:
            print("Нет вакансий для отображения")
            return

        # Сортировка по зарплате через lambda
        sorted_vacancies = sorted(self._vacancies, key=lambda v: v.get_average_salary(), reverse=True)

        for i, v in enumerate(sorted_vacancies[:10], 1):
            avg_salary = v.get_average_salary()
            print(f"{i}. {v.title} @ {v.company}")
            print(f"   Зарплата: {avg_salary:,.0f} руб. | {v.city} | {v.experience_level}")
            print(f"   Навыки: {', '.join(v.skills[:3])}")
            print()

        if len(sorted_vacancies) > 10:
            print(f"... и ещё {len(sorted_vacancies) - 10} вакансий")

    def _parse_vacancies(self):
        """Парсинг новых вакансий"""
        print("\n=== ПАРСИНГ ВАКАНСИЙ ===")

        query = input("Поисковый запрос (по умолчанию Python): ").strip() or "Python"
        city = input("Город (по умолчанию Москва): ").strip() or "Москва"

        # Проверка robots.txt
        parser = JobParser()
        if parser.check_robots_txt("https://hh.ru"):
            print("✅ robots.txt разрешает парсинг")

        new_vacancies = parser.parse_hh_vacancies(query, city)
        self._vacancies.extend(new_vacancies)
        self._analytics = Analytics(self._vacancies, self._tracker.get_all_applications())
        print(f"✅ Добавлено {len(new_vacancies)} новых вакансий")

    def _apply_to_vacancy(self):
        """Подать отклик на вакансию"""
        print("\n=== ПОДАТЬ ОТКЛИК ===")

        self._show_vacancies()

        try:
            idx = int(input("\nНомер вакансии для отклика: ")) - 1
            if 0 <= idx < len(self._vacancies):
                vacancy = self._vacancies[idx]
                notes = input("Заметки (необязательно): ").strip()

                application = self._tracker.add_application(vacancy, notes)
                self._analytics = Analytics(self._vacancies, self._tracker.get_all_applications())
                print(f"✅ Отклик {application.application_id} подан на вакансию {vacancy.title}")
                self._log_operation("APPLY", f"Applied to {vacancy.title} at {vacancy.company}")
            else:
                print("⚠️ Неверный номер вакансии")
        except ValueError:
            print("⚠️ Введите число")

    def _show_applications(self):
        """Показать все отклики"""
        print("\n=== МОИ ОТКЛИКИ ===")

        apps = self._tracker.get_all_applications()

        if not apps:
            print("Нет откликов")
            return

        # Группировка по статусам
        for status in Application.VALID_STATUSES:
            status_apps = list(filter(lambda a: a.status == status, apps))
            if status_apps:
                print(f"\n📌 {status.upper()} ({len(status_apps)}):")
                for app in status_apps:
                    days = app.get_waiting_days()
                    print(f"   {app.application_id}: {app.vacancy.title} @ {app.vacancy.company} ({days} дн.)")

    def _update_application_status(self):
        """Обновить статус отклика"""
        print("\n=== ОБНОВИТЬ СТАТУС ОТКЛИКА ===")

        apps = self._tracker.get_all_applications()
        for i, app in enumerate(apps, 1):
            print(f"{i}. {app.application_id}: {app.vacancy.title} - {app.status}")

        try:
            idx = int(input("\nНомер отклика: ")) - 1
            if 0 <= idx < len(apps):
                app = apps[idx]
                print(f"Текущий статус: {app.status}")
                print(f"Доступные статусы: {Application.VALID_STATUSES}")
                new_status = input("Новый статус: ").strip()

                if new_status in Application.VALID_STATUSES:
                    reason = ""
                    if new_status == 'rejected':
                        reason = input("Причина отказа: ").strip()
                        app.reject(reason)
                    else:
                        app.update_status(new_status)

                    self._analytics = Analytics(self._vacancies, self._tracker.get_all_applications())
                    print(f"✅ Статус обновлён на {new_status}")
                    self._log_operation("STATUS_UPDATE", f"Application {app.application_id} -> {new_status}")
                else:
                    print("⚠️ Недопустимый статус")
            else:
                print("⚠️ Неверный номер отклика")
        except ValueError:
            print("⚠️ Введите число")

    def _show_analytics(self):
        """Показать аналитику"""
        print("\n=== АНАЛИТИКА РЫНКА ===")

        if self._analytics:
            print(self._analytics.generate_report())
        else:
            print("Нет данных для аналитики")

    def _export_to_excel(self):
        """Экспорт данных в Excel"""
        print("\n=== ЭКСПОРТ В EXCEL ===")

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()

            # Лист с вакансиями
            ws_vacancies = wb.active
            ws_vacancies.title = "Вакансии"

            headers = ["ID", "Должность", "Компания", "Зарплата мин", "Зарплата макс",
                      "Средняя", "Город", "Тип занятости", "Опыт", "Навыки"]

            # Жирные заголовки
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for col, header in enumerate(headers, 1):
                cell = ws_vacancies.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                ws_vacancies.column_dimensions[chr(64 + col)].width = 15

            for row, vacancy in enumerate(self._vacancies, 2):
                ws_vacancies.cell(row=row, column=1, value=vacancy.vacancy_id)
                ws_vacancies.cell(row=row, column=2, value=vacancy.title)
                ws_vacancies.cell(row=row, column=3, value=vacancy.company)
                ws_vacancies.cell(row=row, column=4, value=vacancy.salary_min)
                ws_vacancies.cell(row=row, column=5, value=vacancy.salary_max)
                ws_vacancies.cell(row=row, column=6, value=vacancy.get_average_salary())
                ws_vacancies.cell(row=row, column=7, value=vacancy.city)
                ws_vacancies.cell(row=row, column=8, value=vacancy.employment_type)
                ws_vacancies.cell(row=row, column=9, value=vacancy.experience_level)
                ws_vacancies.cell(row=row, column=10, value=", ".join(vacancy.skills))

            # Лист с откликами
            ws_apps = wb.create_sheet("Отклики")
            app_headers = ["ID отклика", "Вакансия", "Компания", "Статус", "Дата отклика",
                          "Дней ожидания", "Заметки"]

            for col, header in enumerate(app_headers, 1):
                cell = ws_apps.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                ws_apps.column_dimensions[chr(64 + col)].width = 18

            for row, app in enumerate(self._tracker.get_all_applications(), 2):
                ws_apps.cell(row=row, column=1, value=app.application_id)
                ws_apps.cell(row=row, column=2, value=app.vacancy.title)
                ws_apps.cell(row=row, column=3, value=app.vacancy.company)
                ws_apps.cell(row=row, column=4, value=app.status)
                ws_apps.cell(row=row, column=5, value=app.applied_date[:10])
                ws_apps.cell(row=row, column=6, value=app.get_waiting_days())
                ws_apps.cell(row=row, column=7, value=app.notes)

            # Лист с аналитикой
            ws_analytics = wb.create_sheet("Аналитика")

            if self._analytics:
                stats = self._analytics.get_salary_statistics()
                ws_analytics.cell(row=1, column=1, value="Метрика")
                ws_analytics.cell(row=1, column=2, value="Значение")
                ws_analytics.cell(row=1, column=1).font = header_font
                ws_analytics.cell(row=1, column=2).font = header_font

                metrics = [
                    ("Средняя зарплата", stats['average']),
                    ("Медианная зарплата", stats['median']),
                    ("Минимальная зарплата", stats['min']),
                    ("Максимальная зарплата", stats['max']),
                    ("Всего вакансий", stats['total_vacancies'])
                ]

                for row, (metric, value) in enumerate(metrics, 2):
                    ws_analytics.cell(row=row, column=1, value=metric)
                    ws_analytics.cell(row=row, column=2, value=value)

            filename = os.path.join(self.DATA_DIR, "analytics_report.xlsx")
            wb.save(filename)
            print(f"✅ Отчёт экспортирован в {filename}")
            self._log_operation("EXPORT_EXCEL", f"Exported to {filename}")

        except ImportError:
            print("❌ Модуль openpyxl не установлен. Установите: pip install openpyxl")
        except Exception as e:
            print(f"❌ Ошибка экспорта: {e}")

    def run(self):
        """Запуск приложения"""
        print("\n🎯 Добро пожаловать в Агрегатор вакансий и Трекер поиска работы!")

        while True:
            print(self._show_menu())
            choice = input("Выберите пункт меню: ").strip()

            if choice == '1':
                self._show_vacancies()
            elif choice == '2':
                self._filter_vacancies()
            elif choice == '3':
                self._parse_vacancies()
            elif choice == '4':
                self._apply_to_vacancy()
            elif choice == '5':
                self._show_applications()
            elif choice == '6':
                self._update_application_status()
            elif choice == '7':
                self._show_analytics()
            elif choice == '8':
                self._export_to_excel()
            elif choice == '9':
                self._save_data()
            elif choice == '0':
                self._save_data()
                print("\n👋 До свидания! Удачи в поиске работы!")
                break
            else:
                print("⚠️ Неверный выбор. Попробуйте снова.")

            input("\nНажмите Enter чтобы продолжить...")


if __name__ == "__main__":
    app = JobTrackerApp()
    app.run()
